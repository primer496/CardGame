"""
xlsx_to_json.py
Converts GameConfig.xlsx sheets to JSON files in Assets/Resources/Config/.

Sheet format:
  Column 1 = field name (camelCase, matches C# SO field name)
  Column 2 = value (string or number)
  Column 3 = comment (ignored)
  Row 1     = header (skipped)

Usage:
  python xlsx_to_json.py
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("[ERROR] openpyxl not found: pip install openpyxl")
    sys.exit(1)

# -- Path config --
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH   = os.path.join(SCRIPT_DIR, "Assets", "Resources", "GameConfig.xlsx")
OUTPUT_DIR  = os.path.join(SCRIPT_DIR, "Assets", "Resources", "Config")

# Sheet name -> value column type ('str' or 'num')
SHEET_VALUE_TYPES = {
    "AudioConfig":   "str",
    "DeckConfig":    "num",
    "CardGameRule":  "num",
}


def sheet_to_dict(ws, value_type: str) -> dict:
    """Convert worksheet rows to {fieldName: value} dict. Skips header and blank rows."""
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        field = row[0]
        raw   = row[1]
        if not field or raw is None:
            continue
        field = str(field).strip()
        if value_type == "num":
            try:
                result[field] = int(raw) if float(raw) == int(float(raw)) else float(raw)
            except (ValueError, TypeError):
                result[field] = raw
        else:
            result[field] = str(raw).strip()
    return result


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"[ERROR] Excel file not found: {XLSX_PATH}")
        sys.exit(1)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    changed = []

    for sheet_name, value_type in SHEET_VALUE_TYPES.items():
        if sheet_name not in wb.sheetnames:
            print(f"[WARN]  Sheet not found, skipping: {sheet_name}")
            continue

        data = sheet_to_dict(wb[sheet_name], value_type)
        out_path = os.path.join(OUTPUT_DIR, f"{sheet_name}.json")

        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        changed.append(out_path)
        print(f"[OK]    {sheet_name} -> {out_path}  ({len(data)} fields)")

    wb.close()

    if changed:
        print(f"\n{len(changed)} JSON file(s) written.")
        print("In Unity Editor, run menu: CardGame -> Import Config from JSON")
    else:
        print("[WARN] No files written. Check that sheet names match SHEET_VALUE_TYPES keys.")


if __name__ == "__main__":
    main()
